import argparse
from date_utils import *
from copy import deepcopy

parser = argparse.ArgumentParser(description='Clean a file containing dates (with column headers containing "date")')
parser.add_argument('file', action='store', type=str, help='path to file')
parser.add_argument('output_file', action='store', type=str, help='path to output file')
parser.add_argument('--sheet_name', dest='sheet_name', nargs=1, action='store', type=str, help='name of worksheet sheet', default=['%DEFAULT_PARAM%'])
parser.add_argument('--init_date', dest='init_date', nargs=1, action='store', type=str, help='initial date as string. Please do not enter an ambiguous date. Default 2020-Feb-01', default=['2020-Feb-01'])
parser.add_argument('--fix_mac', dest='fix_mac', nargs=1, action='store', type=int, help='Whether or not to attempt to fix pre-mac 2011 dates. 0 (False) or 1 (True), default 1', default=[1])
parser.add_argument('--infer_cols', dest='infer_cols', nargs=1, action='store', type=int, help='Whether or not to attempt to infer dates from nearby dates in the same column.  0 (False) or 1 (True), default 1', default=[1])
parser.add_argument('--infer_rows', dest='infer_rows', nargs=1, action='store', type=int, help='Whether or not to attempt to infer dates from nearby dates in the same row.  0 (False) or 1 (True), default 1', default=[1])
parser.add_argument('--searchrx', dest='searchrx', nargs=1, action='store', type=int, help='Search radius for date inferencing from same column. Default 7', default=[7])
parser.add_argument('--skiprows', dest='skiprows', nargs=1, action='store', type=int, help='Number of rows to skip at top of excel file. Default 0', default=[0])
parser.add_argument('--verbosity', dest='verbosity', nargs=1, action='store', type=int, help='How verbose the explanation columns should be. 0 to disable, 1 to output last explanation, 2 for full explanation', default=[1])

args = parser.parse_args()

infile = args.file
print('Reading excel file... ')

sheet_name = args.sheet_name[0]
if sheet_name == '%DEFAULT_PARAM%':
    print("No sheet name input; defaulting to first sheet in workbook.")
    sheet_name = 0
sheet_0 = pd.read_excel(infile, sheet_name=sheet_name, skiprows=args.skiprows[0])

start_date = pd.to_datetime(args.init_date[0])
end_date = pd.to_datetime(getmtime(infile), unit='s') # Time of last modification
explanation_sheet = pd.DataFrame()

if args.fix_mac[0]:
    print('Casting mac-pre2011 dates to modern dates')
    for col_name in sheet_0: 
        if col_name.lower().find('date') != -1:
            print('Working on column: ' + col_name)
            recasted = []
            explanation_col = []
            for date in sheet_0[col_name]:
                explained_date, explanation = date_recast(date, start_date, end_date)
                recasted.append(explained_date)
                explanation_col.append(explanation)
            sheet_0[col_name] = recasted
            explanation_sheet[col_name] = explanation_col
    unambiguous_sheet = sheet_0

if args.infer_cols[0] or args.infer_rows[0]:
    print('Inferring columns or rows.')

    print('Getting unambiguous dates.')
    unambiguous_sheet = pd.DataFrame()
    for col in sheet_0.columns:
        if col.lower().find('date') != -1:
            unambiguous_column, new_explanation_column = get_unambiguous_date_values(sheet_0[col], start_date, end_date)
            unambiguous_sheet[col] = unambiguous_column
            explanation_sheet[col] = [existing_explanation + '|' + new_explanation_column[i] for i, existing_explanation in enumerate(explanation_sheet[col])] 
        else: 
            unambiguous_sheet[col] = sheet_0[col]

    if args.infer_cols[0]:
        print('Inferring columns.')
        for col in sheet_0.columns:
            i = 0
            if col.lower().find('date') != -1:
                last_col_version = deepcopy(unambiguous_sheet[col]).tolist()
                new_col, new_explanation_column = line_fill_ambiguous_date(unambiguous_sheet[col], sheet_0[col], args.searchrx[0])
                explanation_sheet[col] = [existing_explanation + '|' + new_explanation_column[i] for i, existing_explanation in enumerate(explanation_sheet[col])] 

                if new_col != last_col_version:
                    i += 1
                
                while not new_col == last_col_version:
                    last_col_version = deepcopy(new_col)
                    new_col, new_explanation_column = line_fill_ambiguous_date(new_col, sheet_0[col], args.searchrx[0]) 
                    explanation_sheet[col] = [existing_explanation + '|' + new_explanation_column[i] for i, existing_explanation in enumerate(explanation_sheet[col])] 

                    i += 1
                    if i > 100:
                        break
                unambiguous_sheet[col] = new_col
                print(col +' has been modified '+str(i) + ' times by iterative column fill method')

    if args.infer_rows[0]:
        print('Inferring rows.')
        unambiguous_sheet, new_explanation_sheet = row_fill_ambiguous_date(unambiguous_sheet, sheet_0)
        for col in explanation_sheet.columns:
            explanation_sheet[col] = [existing_explanation + '|' + new_explanation_sheet[col][i] for i, existing_explanation in enumerate(explanation_sheet[col])]

else:
    print('Skipping column/row inference.')

if args.verbosity[0] == 0:
    unambiguous_sheet.to_excel(args.output_file, index=False)
else:
    completed_df = pd.DataFrame()
    for col in unambiguous_sheet.columns:
        completed_df[col] = unambiguous_sheet[col]
        if col in explanation_sheet.columns:
            if args.verbosity[0] == 1:
                column_explanation = [list(filter(lambda x: x != '', x.split('|'))) for x in explanation_sheet[col]]
                column_explanation = [(lambda x: x[-1] if len(x) > 0 else "")(x) for x in column_explanation]
            elif args.verbosity[0] == 2:
                column_explanation = explanation_sheet[col]

            completed_df[col+'_explanations'] = column_explanation
    completed_df = completed_df.fillna("NA")

    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import PatternFill
    wb = Workbook()
    ws = wb.active

    for r in dataframe_to_rows(completed_df, header=True, index=False):
        ws.append(r)
    
    for column in ws.columns:
        if column[0].value is not None:
            if 'date' in column[0].value.lower() and 'explanations' not in column[0].value.lower():
                for cell in column:
                    cell.fill = PatternFill("solid", fgColor="DDDDDD")
    wb.save(args.output_file)
    #completed_df.to_excel(args.output_file, index=False)





